#!/usr/bin/env python
"""Flatten portfolio.json into a formatted Google Sheets-compatible .xlsx workbook.

Requires: openpyxl (pip install openpyxl)
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Install openpyxl: pip install openpyxl")

HERE = os.path.dirname(os.path.abspath(__file__))

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2B2B2B", end_color="2B2B2B", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="left", vertical="center", wrap_text=True)

GEO_FILL = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
REGION_FILL = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
TERRITORY_FILL = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")

ACT_FILL = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
WATCH_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

WRAP_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
TOP_ALIGNMENT = Alignment(vertical="top")

# ---------------------------------------------------------------------------
# Column definitions: (name, width)
# ---------------------------------------------------------------------------
PORTFOLIO_COLUMNS = [
    ("Level", 12),
    ("Name", 35),
    ("Parent", 25),
    ("Accounts", 10),
    ("Signal Score", 14),
    ("Total Activities", 16),
    ("Meetings 30d", 14),
    ("Emails 30d", 13),
    ("Opportunities", 14),
    ("Activity Trend", 14),
    ("Signals ACT", 14),
    ("Signals WATCH", 15),
    ("Top Signal", 50),
    ("Top Action Item", 50),
    ("Match Status", 13),
    ("Summary", 60),
    ("Recommended Next Move", 50),
]

SIGNALS_COLUMNS = [
    ("Account", 30),
    ("Disposition", 12),
    ("Score", 8),
    ("Headline", 50),
    ("What Changed", 45),
    ("Why It Matters", 45),
    ("Red Hat Relevance", 45),
    ("Recommended Action", 45),
    ("Source Type", 16),
    ("Confidence", 12),
    ("Source URL", 20),
    ("Published", 12),
]


# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------

def load_portfolio(path):
    """Read JSON and validate required top-level keys."""
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    for key in ("run", "scope", "summary", "accounts"):
        if key not in data:
            sys.exit(f"portfolio.json missing required key: {key}")
    return data


def safe_str(value):
    """None->empty string, list->newline-joined, else str()."""
    if value is None:
        return ""
    if isinstance(value, list):
        return "\n".join(str(v) for v in value)
    return str(value)


def extract_metric(account, key, default=None):
    """Safely access account['internal']['metrics'][key]."""
    try:
        return account["internal"]["metrics"][key]
    except (KeyError, TypeError):
        return default


def top_signal(accounts):
    """Find the highest-scored signal across a list of account dicts.

    Returns the signal dict or None.
    """
    best = None
    best_score = None
    for acct in accounts:
        for sig in acct.get("signals") or []:
            score = sig.get("score")
            if score is not None and (best_score is None or score > best_score):
                best = sig
                best_score = score
    return best


def aggregate_accounts(accounts):
    """Sum up numeric metrics from a list of accounts."""
    account_count = len(accounts)
    total_activities = 0
    meetings_30d = 0
    emails_30d = 0
    opportunities = 0
    act_count = 0
    watch_count = 0

    for acct in accounts:
        total_activities += (extract_metric(acct, "total_activities", 0) or 0)
        meetings_30d += (extract_metric(acct, "meeting_count_30d", 0) or 0)
        emails_30d += (extract_metric(acct, "email_count_30d", 0) or 0)

        opp_names = extract_metric(acct, "linked_opportunity_names")
        opp_count = extract_metric(acct, "linked_opportunity_count", 0)
        if isinstance(opp_names, list):
            opportunities += len(opp_names)
        else:
            opportunities += (opp_count or 0)

        for sig in acct.get("signals") or []:
            disp = sig.get("disposition", "")
            if disp == "ACT":
                act_count += 1
            elif disp == "WATCH":
                watch_count += 1

    signal_scores = [a.get("signal_score") for a in accounts
                     if a.get("signal_score") is not None]
    avg_signal_score = (round(sum(signal_scores) / len(signal_scores))
                        if signal_scores else None)

    ts = top_signal(accounts)

    return {
        "account_count": account_count,
        "signal_score": avg_signal_score,
        "total_activities": total_activities,
        "meetings_30d": meetings_30d,
        "emails_30d": emails_30d,
        "opportunities": opportunities,
        "act_count": act_count,
        "watch_count": watch_count,
        "top_signal_headline": ts.get("headline") if ts else None,
        "top_action": ts.get("recommended_action") if ts else None,
    }


def build_hierarchy(accounts):
    """Return nested dict: {geo: {region: {territory: [accounts]}}}."""
    tree = {}
    for acct in accounts:
        h = acct.get("hierarchy", {})
        geo = h.get("geo") or "Unknown"
        region = h.get("region") or "Unknown"
        territory = h.get("territory_name") or "Unknown"
        tree.setdefault(geo, {}).setdefault(region, {}).setdefault(
            territory, []
        ).append(acct)
    return tree


# ---------------------------------------------------------------------------
# Sheet-writing helpers
# ---------------------------------------------------------------------------

def apply_header_style(ws, columns):
    """Write header row with dark fill, white bold font, set column widths,
    freeze panes at A2, and set auto-filter on header row."""
    for col_idx, (name, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"


def _write_portfolio_row(ws, row, values, fill):
    """Write a single row into the Portfolio tab."""
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        if fill is not None:
            cell.fill = fill
        if col_idx >= 13:
            cell.alignment = WRAP_ALIGNMENT
        else:
            cell.alignment = TOP_ALIGNMENT


def write_portfolio_tab(ws, hierarchy, all_accounts):
    """Write the Portfolio tab with GEO, REGION, TERRITORY, and ACCOUNT rows."""
    apply_header_style(ws, PORTFOLIO_COLUMNS)
    num_cols = len(PORTFOLIO_COLUMNS)
    row = 2

    # --- GEO rows ---
    for geo in sorted(hierarchy):
        geo_accounts = []
        for region in hierarchy[geo]:
            for territory in hierarchy[geo][region]:
                geo_accounts.extend(hierarchy[geo][region][territory])
        agg = aggregate_accounts(geo_accounts)
        values = [
            "GEO", geo, "", agg["account_count"], agg["signal_score"],
            agg["total_activities"], agg["meetings_30d"], agg["emails_30d"],
            agg["opportunities"], None, agg["act_count"], agg["watch_count"],
            agg["top_signal_headline"], agg["top_action"],
            None, None, None,
        ]
        _write_portfolio_row(ws, row, values, GEO_FILL)
        row += 1

    # --- REGION rows ---
    for geo in sorted(hierarchy):
        for region in sorted(hierarchy[geo]):
            region_accounts = []
            for territory in hierarchy[geo][region]:
                region_accounts.extend(hierarchy[geo][region][territory])
            agg = aggregate_accounts(region_accounts)
            values = [
                "REGION", region, geo, agg["account_count"], agg["signal_score"],
                agg["total_activities"], agg["meetings_30d"], agg["emails_30d"],
                agg["opportunities"], None, agg["act_count"], agg["watch_count"],
                agg["top_signal_headline"], agg["top_action"],
                None, None, None,
            ]
            _write_portfolio_row(ws, row, values, REGION_FILL)
            row += 1

    # --- TERRITORY rows ---
    for geo in sorted(hierarchy):
        for region in sorted(hierarchy[geo]):
            for territory in sorted(hierarchy[geo][region]):
                territory_accounts = hierarchy[geo][region][territory]
                agg = aggregate_accounts(territory_accounts)
                values = [
                    "TERRITORY", territory, region, agg["account_count"], agg["signal_score"],
                    agg["total_activities"], agg["meetings_30d"], agg["emails_30d"],
                    agg["opportunities"], None, agg["act_count"], agg["watch_count"],
                    agg["top_signal_headline"], agg["top_action"],
                    None, None, None,
                ]
                _write_portfolio_row(ws, row, values, TERRITORY_FILL)
                row += 1

    # --- ACCOUNT rows (sorted by priority descending) ---
    sorted_accounts = sorted(
        all_accounts,
        key=lambda a: (a.get("signal_score") or 0),
        reverse=True,
    )
    for acct in sorted_accounts:
        signals = acct.get("signals") or []
        act_count = sum(1 for s in signals if s.get("disposition") == "ACT")
        watch_count = sum(1 for s in signals if s.get("disposition") == "WATCH")
        ts = top_signal([acct])
        values = [
            "ACCOUNT",
            acct.get("account_name"),
            acct.get("hierarchy", {}).get("territory_name"),
            None,
            acct.get("signal_score"),
            extract_metric(acct, "total_activities"),
            extract_metric(acct, "meeting_count_30d"),
            extract_metric(acct, "email_count_30d"),
            extract_metric(acct, "linked_opportunity_count"),
            extract_metric(acct, "activity_trend"),
            act_count,
            watch_count,
            ts.get("headline") if ts else None,
            ts.get("recommended_action") if ts else None,
            acct.get("identity", {}).get("match_status"),
            safe_str(acct.get("summary")),
            safe_str(acct.get("recommended_next_move")),
        ]
        _write_portfolio_row(ws, row, values, None)
        row += 1

    # Update auto_filter.ref to cover all data rows
    ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}{max(row - 1, 1)}"


def write_signals_tab(ws, accounts):
    """Write the Signals tab with one row per signal across all accounts."""
    apply_header_style(ws, SIGNALS_COLUMNS)
    num_cols = len(SIGNALS_COLUMNS)
    row = 2

    # Collect all signals paired with account_name
    signal_rows = []
    for acct in accounts:
        name = acct.get("account_name", "")
        for sig in acct.get("signals") or []:
            signal_rows.append((name, sig))

    # Sort by account_name ascending, then score descending
    signal_rows.sort(key=lambda x: (x[0], -(x[1].get("score") or 0)))

    for acct_name, sig in signal_rows:
        disposition = sig.get("disposition", "")
        source_url = sig.get("source_url")

        values = [
            acct_name,
            disposition,
            sig.get("score"),
            sig.get("headline"),
            sig.get("what_changed"),
            sig.get("why_it_matters"),
            sig.get("red_hat_relevance"),
            sig.get("recommended_action"),
            sig.get("source_type"),
            sig.get("confidence"),
            None,  # source_url handled separately below
            sig.get("published_at"),
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            if col_idx >= 4:
                cell.alignment = WRAP_ALIGNMENT
            else:
                cell.alignment = TOP_ALIGNMENT

        # Disposition cell fill
        disp_cell = ws.cell(row=row, column=2)
        if disposition == "ACT":
            disp_cell.fill = ACT_FILL
        elif disposition == "WATCH":
            disp_cell.fill = WATCH_FILL

        # Source URL cell — hyperlink if available
        if source_url:
            url_cell = ws.cell(row=row, column=11)
            url_cell.hyperlink = source_url
            url_cell.value = "Open"
            url_cell.font = Font(color="0563C1", underline="single")

        row += 1

    # Update auto_filter.ref to cover all data rows
    ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}{max(row - 1, 1)}"


# ---------------------------------------------------------------------------
# Output path derivation
# ---------------------------------------------------------------------------

def derive_output_path(portfolio, input_path):
    """Derive default output path from scope.value."""
    scope_value = portfolio.get("scope", {}).get("value", "portfolio")
    slug = re.sub(r"[^a-zA-Z0-9]+", "_", scope_value).strip("_")[:40]
    directory = os.path.dirname(input_path)
    if not directory:
        directory = "."
    return os.path.join(directory, f"{slug}.xlsx")


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("portfolio", help="Path to portfolio.json")
    parser.add_argument("--out", help="Output .xlsx path (default: derived from scope)")
    args = parser.parse_args()

    portfolio = load_portfolio(args.portfolio)
    accounts = portfolio["accounts"]

    out_path = args.out or derive_output_path(portfolio, args.portfolio)

    wb = Workbook()

    # Portfolio tab (default active sheet)
    ws_portfolio = wb.active
    ws_portfolio.title = "Portfolio"
    hierarchy = build_hierarchy(accounts)
    write_portfolio_tab(ws_portfolio, hierarchy, accounts)

    # Signals tab
    ws_signals = wb.create_sheet("Signals")
    write_signals_tab(ws_signals, accounts)

    wb.save(out_path)

    signal_count = sum(len(a.get("signals") or []) for a in accounts)
    print(
        f"Wrote {out_path} — {len(accounts)} accounts, {signal_count} signals",
        file=sys.stderr,
    )


if __name__ == "__main__":
    main()
